#library to import the excel file
import openpyxl
import shutil

from sendEmail import sendEmail



#libraries to create the pdf file and add text to it
from reportlab.pdfgen import canvas

from reportlab.lib.styles import (ParagraphStyle, getSampleStyleSheet)
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.pdfmetrics import stringWidth
from reportlab.pdfbase.ttfonts import TTFont
#libraries to merge pdf files

from reportlab.lib.units import inch, cm
import os
from PyPDF2 import PdfFileReader, PdfFileMerger, PdfFileWriter

from reportlab.lib import pdfencrypt


from reportlab.platypus import Table, TableStyle, Paragraph

from reportlab.platypus import SimpleDocTemplate, Spacer, Image
from reportlab.lib.pagesizes import A4, A5, letter

from reportlab.platypus import TableStyle
from reportlab.lib import colors

from flask import Flask, render_template, request, jsonify, send_file
from zipfile import ZipFile
import os
import requests

app = Flask(__name__)

@app.route('/')
def index():
  return render_template("index.html")

@app.route('/extract', methods=["POST"])

def create_payslip():

    zipObj = ZipFile('sample.zip', 'w')

    #this was changed
    #convert the font so it is compatible
    pdfmetrics.registerFont(TTFont('Arial','arial.ttf'))


    r = request.files['excel']

    #import the sheet from the excel file
    wb = openpyxl.load_workbook(r, data_only=True)
    sheet = wb[wb.sheetnames[0]]

    #print(sheet.cell(4,1).value)
    #Page information
    page_width = 2156
    page_height = 3050
    spread = 50
    start = 200
    start_2 = 700

    #Payslip variables
    company_name = 'Conacent Consulting'

    logo = Image("logo and address.png")
    logo._restrictSize(6*inch,8*inch)
    logo.hAlign = "LEFT"
    logo.VALIGN = "TOP"
    col_names = []
    for name in sheet[2]:
        if name.value and name.value != "email":
            col_names.append(name.value)
        

    super_columns = []
    for super_column in sheet[1]:
        super_columns.append(super_column.value)

    year = ""

    styles = getSampleStyleSheet()
    i = 3
    emails = []
    while (i):
        vals = []
        password = ""
        if sheet.cell(row = i, column = 1).value is None:
            break
        for j in range(1,sheet.max_column+1):
           
            if sheet.cell(row = i, column = j).value is not None:
                inp = str(sheet.cell(row = i, column = j).value)
        
                if "00:00:00" in inp:

                    inp = inp.replace("00:00:00","")    
                      
                    y = inp[:inp.find("-")]
                    # inp = inp[::-1]
                    month = inp[inp.find("-") + 1: inp.rfind("-")]
                    date = inp[inp.rfind("-") + 1: ]
                    inp = date + "/" + month + "/" + y
                    inp = inp.replace(" ","")             
                if ("#" in inp):
                    print("this date of joining is weird")
                if "@" in inp:
                    emails.append(inp)
                else :
                    vals.append(inp)
            else:
                vals.append("N/A")
        name =  str(vals[1])+ ' ' + str(vals[0])  + '.pdf' 
        email = emails[i - 3]
        print(email + " this is email")
        pdf = SimpleDocTemplate(
                    name,
                    pagesize= A4,
                    )

        if str(vals[2]) != 'x' and str(vals[2]) != 'N/A' and type(vals[0]) == str:
            data = []
            elements = [logo]
            for j in range(len(vals)):
                if super_columns[j] is None:
                    if str(vals[j]) == 'N/A':
                        continue
                    data.append([col_names[j], str(vals[j])])
                elif super_columns[j] == "LEAVE STATEMENT":
                    table = Table(data, rowHeights = len(data) * [13], colWidths=inch*3)
                    style = TableStyle([
                            ('BACKGROUND', (0,0), (3,0), colors.steelblue),
                            ('TEXTCOLOR',(0,0),(-1,0),colors.whitesmoke),

                            ('ALIGN',(0,0),(-1,-1),'CENTER'),

                            ('FONTNAME', (0,0), (-1,0), 'Courier-Bold'),
                            ('FONTSIZE', (0,0), (-1,0), 10),

                            ('BOTTOMPADDING', (0,0), (-1,0), 2),

                            ('BACKGROUND',(0,1),(-1,-1),colors.beige),
                        ])
                    table.hAlign = "LEFT"
                    
                    table.setStyle(style)


                    elements.append(table)
                    data = []
                    data.append([super_columns[j], ""])
                    data.append([col_names[j], str(vals[j])])
                else:
                    if len(data) == 1:
                        continue
                    #This changes main table
                    #13 IS BEST FONT SIZE
                    table = Table(data, rowHeights = len(data) * [13], colWidths=inch*3)
                    if data[0][1] != "":

                        #First part
                       
                        data[0] = [Paragraph("<b> SALARY FOR" + " " + data[0][0].upper() +  "</b>")]
                        table = Table(data)
                        style = TableStyle([
                        ('BACKGROUND', (0,0), (3,0), colors.white),
                        ('TEXTCOLOR',(0,0),(-1,0),colors.black),

                        ('ALIGN',(0,0),(-1,-1),'LEFT'),

                        ('VALIGN',(-200,-200),(-100,-100),'TOP'),

                        ('FONTNAME', (0,0), (-1,0), 'Courier-Bold'),
                        ('FONTSIZE', (0,0), (-1,0), 100),
                        ('TEXTFONT', (0, 1), (-1, 1), 'Times-Bold'),

                        ('BOTTOMPADDING', (0,0), (-1,0), 2),

                        ('BACKGROUND',(0,1),(-1,-1),colors.white),
                        ])
                        table.hAlign = "LEFT"
                    else:         
                        
                        style = TableStyle([
                            ('BACKGROUND', (0,0), (3,0), colors.steelblue),
                            ('TEXTCOLOR',(0,0),(-1,0),colors.whitesmoke),

                            ('ALIGN',(0,0),(-1,-1),'CENTER'),

                            ('FONTNAME', (0,0), (-1,0), 'Courier-Bold'),
                            #('FONTSIZE', (0,0), (-1,0), 12),
                            ('TEXTSIZE', (0,0), (-1,0), 10),
                            #('TOPPADDING', (0,0), (-1,0), 2),
                            ('BOTTOMPADDING', (0,0), (-1,0), 2),

                            ('BACKGROUND',(0,1),(-1,-1),colors.beige),
                        ])
                        table.hAlign = "LEFT"
                    table.setStyle(style)
                    elements.append(table)
                    elements.append(Spacer(1,5))
                    data = []
                    data.append([super_columns[j], ""])
                    elements.append(Spacer(1,10))
                    data.append([col_names[j], vals[j]])
            
            #this is for leave 
            elements.append(Spacer(1,5))
            table = Table(data, rowHeights = len(data) * [13], colWidths=inch*3)
            style = TableStyle([
                                        ('BACKGROUND', (0,0), (3,0), colors.steelblue),
                                        ('TEXTCOLOR',(0,0),(-1,0),colors.whitesmoke),

                                        ('ALIGN',(0,0),(-1,-1),'CENTER'),

                                        ('FONTNAME', (0,0), (-1,0), 'Courier-Bold'),
                                        ('BOTTOMPADDING', (0,0), (-1,0), 2),

                                        ('BACKGROUND',(0,1),(-1,-1),colors.beige),
                                    ])
            table.hAlign = "LEFT"
            table.setStyle(style)
            elements.append(Spacer(1,10))
            elements.append(table)
            elements.append(Spacer(1,120))
            style_new = getSampleStyleSheet()
            yourStyle = ParagraphStyle('yourtitle',
                           fontName="Helvetica",
                           fontSize=8,
                           parent=style_new['Heading2'],
                           alignment=1,
                           spaceAfter=2)
            elements.append(Paragraph(("<i>Signature not required for this payslip </i>"), yourStyle))
            elements.append(Paragraph("<i> Registered Office:  P-94/95, Bangur Avenue, BL-C, Kolkata - 700055 </i>",  yourStyle))
            pdf.build(elements)

            # create a PdfFileWriter object
            out = PdfFileWriter()

            
            # Open our PDF file with the PdfFileReader
            filename = PdfFileReader(name)
            out.appendPagesFromReader(filename)
            
            if password == "":
                password = vals[5]

            out.encrypt(user_pwd = password)


            with open(name, "wb") as f:
                # Write our encrypted PDF to this file
                out.write(f)
            #return send_file(name)

            #zipObj.write(name)


            i += 1
            year = vals[0]
            if (year != "N/A"):
                import os 
                if os.path.exists(name):
                    sendEmail(name, email, month=str(vals[0]), personName= str (vals[1]))
                    os.remove(name)
                else:
                    print("The file does not exist")

          
    # if (year != "N/A"):
    #     merge_pdfs(year)
            #Saving the pdf file
    zipObj.close()
    return "Salary slips have been sent", 200
    # return send_file('sample.zip')
        
def merge_pdfs(year):

    files_dir = 'C:\\Users\\tejes\Desktop\conacent\payslips' 
    SOURCE_DIR = 'C:\\Users\\tejes\Desktop\conacent\payslips' 




    import os 
    from os import path
    # Directory 
        
    # Parent Directory path 
        
    DEST_DIR = 'C:\\Users\\tejes\Desktop\conacent\payslips\payslipsFolder ' + year 
    # Path 
    p = os.path.normpath(DEST_DIR)
    if path.exists(p):
        shutil.rmtree(p, ignore_errors = False) 
    os.mkdir(p) 


    pdf_files = [f for f in os.listdir(files_dir) if f.endswith('.pdf')] #Get all files in the directory that end with '.pdf'
    merger = PdfFileMerger() #Create an empty file
    for fname in pdf_files:

        #merger.append(PdfFileReader(os.path.join(files_dir,filename),'rb')) #Add every pdf to the empty file
        #erger.write(PdfFileReader(os.path.join('C:\\Users\\tejes\Desktop\conacent\payslipsFolder',filename))) #Save the file
        shutil.move(os.path.join(SOURCE_DIR, fname), DEST_DIR)
    
    shutil.rmtree(p, ignore_errors = False) 

#create_payslip()
#merge_pdfs()

if __name__ == "__main__":
    app.run(debug = True, threaded=True, port = int(os.environ.get('PORT', 5000)))